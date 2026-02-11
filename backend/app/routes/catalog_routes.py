"""
Catalog Routes
Gestione cataloghi condivisibili di asset
"""

from flask import Blueprint, request, jsonify, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from app import db
from app.models import Catalog, InventoryAsset, InventoryCategory, Club
from datetime import datetime
import secrets
import json

catalog_bp = Blueprint('catalog', __name__)


def verify_club():
    """Verifica che l'utente sia un club"""
    claims = get_jwt()
    if claims.get('role') != 'club':
        return None
    club_id = int(get_jwt_identity())
    return club_id


def generate_public_token():
    """Genera un token univoco per il catalogo pubblico"""
    while True:
        token = secrets.token_urlsafe(32)
        if not Catalog.query.filter_by(public_token=token).first():
            return token


# =============================================================================
# CRUD CATALOGHI
# =============================================================================

@catalog_bp.route('/catalogs', methods=['GET'])
@jwt_required()
def get_catalogs():
    """Lista cataloghi del club"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalogs = Catalog.query.filter_by(club_id=club_id).order_by(Catalog.created_at.desc()).all()

    return jsonify({
        'catalogs': [c.to_dict() for c in catalogs]
    }), 200


@catalog_bp.route('/catalogs', methods=['POST'])
@jwt_required()
def create_catalog():
    """Crea nuovo catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    data = request.json

    # Genera token pubblico
    public_token = generate_public_token()

    # Gestisce layout_json - può essere dict o stringa
    layout_json = data.get('layout_json')
    if layout_json and isinstance(layout_json, dict):
        layout_json = json.dumps(layout_json)

    # Crea catalogo
    catalog = Catalog(
        club_id=club_id,
        nome=data.get('nome', 'Nuovo Catalogo'),
        descrizione=data.get('descrizione'),
        public_token=public_token,
        logo_url=data.get('logo_url'),
        colore_primario=data.get('colore_primario'),
        colore_secondario=data.get('colore_secondario'),
        immagine_copertina=data.get('immagine_copertina'),
        layout_json=layout_json,
        mostra_prezzi=data.get('mostra_prezzi', True),
        mostra_disponibilita=data.get('mostra_disponibilita', True),
        messaggio_benvenuto=data.get('messaggio_benvenuto'),
        messaggio_footer=data.get('messaggio_footer'),
        email_contatto=data.get('email_contatto'),
        telefono_contatto=data.get('telefono_contatto'),
        attivo=data.get('attivo', True)
    )

    # Aggiungi asset se specificati
    asset_ids = data.get('asset_ids', [])
    if asset_ids:
        assets = InventoryAsset.query.filter(
            InventoryAsset.id.in_(asset_ids),
            InventoryAsset.club_id == club_id
        ).all()
        catalog.assets = assets

    db.session.add(catalog)
    db.session.commit()

    return jsonify({
        'message': 'Catalogo creato',
        'catalog': catalog.to_dict(include_assets=True)
    }), 201


@catalog_bp.route('/catalogs/<int:catalog_id>', methods=['GET'])
@jwt_required()
def get_catalog(catalog_id):
    """Dettaglio catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    return jsonify({
        'catalog': catalog.to_dict(include_assets=True)
    }), 200


@catalog_bp.route('/catalogs/<int:catalog_id>', methods=['PUT'])
@jwt_required()
def update_catalog(catalog_id):
    """Aggiorna catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    data = request.json

    # Aggiorna campi base
    catalog.nome = data.get('nome', catalog.nome)
    catalog.descrizione = data.get('descrizione', catalog.descrizione)
    catalog.logo_url = data.get('logo_url', catalog.logo_url)
    catalog.colore_primario = data.get('colore_primario', catalog.colore_primario)
    catalog.colore_secondario = data.get('colore_secondario', catalog.colore_secondario)
    catalog.immagine_copertina = data.get('immagine_copertina', catalog.immagine_copertina)

    # Gestisce layout_json - può essere dict o stringa
    if 'layout_json' in data:
        layout_json = data.get('layout_json')
        if layout_json and isinstance(layout_json, dict):
            catalog.layout_json = json.dumps(layout_json)
        else:
            catalog.layout_json = layout_json

    catalog.mostra_prezzi = data.get('mostra_prezzi', catalog.mostra_prezzi)
    catalog.mostra_disponibilita = data.get('mostra_disponibilita', catalog.mostra_disponibilita)
    catalog.messaggio_benvenuto = data.get('messaggio_benvenuto', catalog.messaggio_benvenuto)
    catalog.messaggio_footer = data.get('messaggio_footer', catalog.messaggio_footer)
    catalog.email_contatto = data.get('email_contatto', catalog.email_contatto)
    catalog.telefono_contatto = data.get('telefono_contatto', catalog.telefono_contatto)
    catalog.attivo = data.get('attivo', catalog.attivo)

    # Gestione data scadenza
    if 'data_scadenza' in data:
        if data['data_scadenza']:
            catalog.data_scadenza = datetime.fromisoformat(data['data_scadenza'].replace('Z', '+00:00'))
        else:
            catalog.data_scadenza = None

    # Aggiorna asset se specificati
    if 'asset_ids' in data:
        asset_ids = data['asset_ids']
        assets = InventoryAsset.query.filter(
            InventoryAsset.id.in_(asset_ids),
            InventoryAsset.club_id == club_id
        ).all()
        catalog.assets = assets

    db.session.commit()

    return jsonify({
        'message': 'Catalogo aggiornato',
        'catalog': catalog.to_dict(include_assets=True)
    }), 200


@catalog_bp.route('/catalogs/<int:catalog_id>', methods=['DELETE'])
@jwt_required()
def delete_catalog(catalog_id):
    """Elimina catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    db.session.delete(catalog)
    db.session.commit()

    return jsonify({'message': 'Catalogo eliminato'}), 200


@catalog_bp.route('/catalogs/<int:catalog_id>/regenerate-token', methods=['POST'])
@jwt_required()
def regenerate_token(catalog_id):
    """Rigenera il token pubblico del catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    catalog.public_token = generate_public_token()
    db.session.commit()

    return jsonify({
        'message': 'Token rigenerato',
        'public_token': catalog.public_token
    }), 200


@catalog_bp.route('/catalogs/<int:catalog_id>/duplicate', methods=['POST'])
@jwt_required()
def duplicate_catalog(catalog_id):
    """Duplica un catalogo esistente"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    original = Catalog.query.get(catalog_id)
    if not original or original.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    new_catalog = Catalog(
        club_id=club_id,
        nome=f"{original.nome} (copia)",
        descrizione=original.descrizione,
        public_token=generate_public_token(),
        logo_url=original.logo_url,
        colore_primario=original.colore_primario,
        colore_secondario=original.colore_secondario,
        immagine_copertina=original.immagine_copertina,
        layout_json=original.layout_json,
        mostra_prezzi=original.mostra_prezzi,
        mostra_disponibilita=original.mostra_disponibilita,
        messaggio_benvenuto=original.messaggio_benvenuto,
        messaggio_footer=original.messaggio_footer,
        email_contatto=original.email_contatto,
        telefono_contatto=original.telefono_contatto,
        attivo=True
    )
    new_catalog.assets = list(original.assets)

    db.session.add(new_catalog)
    db.session.commit()

    return jsonify({
        'message': 'Catalogo duplicato',
        'catalog': new_catalog.to_dict(include_assets=True)
    }), 201


# =============================================================================
# CATALOGO PUBBLICO (NO AUTH)
# =============================================================================

@catalog_bp.route('/public/catalog/<token>', methods=['GET'])
def get_public_catalog(token):
    """Visualizza catalogo pubblico tramite token"""
    catalog = Catalog.query.filter_by(public_token=token).first()

    if not catalog:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    if not catalog.attivo:
        return jsonify({'error': 'Catalogo non attivo'}), 403

    # Verifica scadenza
    if catalog.data_scadenza and datetime.utcnow() > catalog.data_scadenza:
        return jsonify({'error': 'Catalogo scaduto'}), 403

    # Incrementa visualizzazioni
    catalog.visualizzazioni += 1
    catalog.ultimo_accesso = datetime.utcnow()
    db.session.commit()

    # Recupera info club
    club = Club.query.get(catalog.club_id)

    # Prepara dati catalogo
    catalog_data = catalog.to_dict(include_assets=True)

    # Aggiungi info club
    catalog_data['club'] = {
        'nome': club.nome,
        'logo_url': club.logo_url,
        'tipologia': club.tipologia,
        'sito_web': club.sito_web,
        'brand_colore_primario': club.brand_colore_primario,
        'brand_colore_secondario': club.brand_colore_secondario
    } if club else None

    # Se non mostra prezzi, rimuovili dagli asset
    if not catalog.mostra_prezzi and 'assets' in catalog_data:
        for asset in catalog_data['assets']:
            asset.pop('prezzo_listino', None)

    # Se non mostra disponibilità, rimuovila dagli asset
    if not catalog.mostra_disponibilita and 'assets' in catalog_data:
        for asset in catalog_data['assets']:
            asset.pop('disponibile', None)
            asset.pop('quantita_disponibile', None)

    return jsonify({'catalog': catalog_data}), 200


# =============================================================================
# EXPORT
# =============================================================================

@catalog_bp.route('/catalogs/<int:catalog_id>/export/excel', methods=['GET'])
@jwt_required()
def export_excel(catalog_id):
    """Esporta catalogo in Excel"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    try:
        import openpyxl
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = catalog.nome[:31]  # Excel limita a 31 caratteri

        # Stili
        header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        headers = ['Nome', 'Categoria', 'Tipo', 'Posizione', 'Dimensioni', 'Prezzo', 'Disponibile']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Dati
        for row, asset in enumerate(catalog.assets, 2):
            data = [
                asset.nome,
                asset.category.nome if asset.category else '',
                asset.tipo,
                asset.posizione or '',
                asset.dimensioni or '',
                f"€ {asset.prezzo_listino:,.2f}" if asset.prezzo_listino else '',
                'Sì' if asset.disponibile else 'No'
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

        # Auto-width colonne
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Salva in buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=catalogo_{catalog.nome}.xlsx'

        return response

    except ImportError:
        return jsonify({'error': 'Modulo openpyxl non installato'}), 500


@catalog_bp.route('/catalogs/<int:catalog_id>/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf(catalog_id):
    """Esporta catalogo in PDF"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    catalog = Catalog.query.get(catalog_id)
    if not catalog or catalog.club_id != club_id:
        return jsonify({'error': 'Catalogo non trovato'}), 404

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        story = []

        # Stile titolo
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1A1A1A')
        )

        # Titolo
        story.append(Paragraph(catalog.nome, title_style))

        # Descrizione
        if catalog.descrizione:
            story.append(Paragraph(catalog.descrizione, styles['Normal']))
            story.append(Spacer(1, 20))

        # Messaggio benvenuto
        if catalog.messaggio_benvenuto:
            story.append(Paragraph(catalog.messaggio_benvenuto, styles['Normal']))
            story.append(Spacer(1, 20))

        # Tabella asset
        if catalog.assets:
            data = [['Nome', 'Categoria', 'Prezzo']]
            for asset in catalog.assets:
                data.append([
                    asset.nome,
                    asset.category.nome if asset.category else '',
                    f"€ {asset.prezzo_listino:,.2f}" if asset.prezzo_listino else '-'
                ])

            table = Table(data, colWidths=[7*cm, 4*cm, 3*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A1A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
            ]))
            story.append(table)

        story.append(Spacer(1, 30))

        # Footer
        if catalog.messaggio_footer:
            story.append(Paragraph(catalog.messaggio_footer, styles['Normal']))

        # Contatti
        if catalog.email_contatto or catalog.telefono_contatto:
            story.append(Spacer(1, 20))
            contact_text = "Contatti: "
            if catalog.email_contatto:
                contact_text += catalog.email_contatto
            if catalog.telefono_contatto:
                contact_text += f" | {catalog.telefono_contatto}"
            story.append(Paragraph(contact_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=catalogo_{catalog.nome}.pdf'

        return response

    except ImportError:
        return jsonify({'error': 'Modulo reportlab non installato'}), 500


# =============================================================================
# ASSET DISPONIBILI
# =============================================================================

@catalog_bp.route('/catalogs/available-assets', methods=['GET'])
@jwt_required()
def get_available_assets():
    """Lista asset disponibili per il catalogo"""
    club_id = verify_club()
    if not club_id:
        return jsonify({'error': 'Accesso non autorizzato'}), 403

    # Recupera tutti gli asset attivi del club
    assets = InventoryAsset.query.filter_by(
        club_id=club_id,
        attivo=True,
        archiviato=False
    ).order_by(InventoryAsset.nome).all()

    # Raggruppa per categoria
    categories = {}
    for asset in assets:
        cat_name = asset.category.nome if asset.category else 'Senza categoria'
        if cat_name not in categories:
            categories[cat_name] = []
        categories[cat_name].append({
            'id': asset.id,
            'nome': asset.nome,
            'descrizione_breve': asset.descrizione_breve,
            'immagine_principale': asset.immagine_principale,
            'prezzo_listino': asset.prezzo_listino,
            'disponibile': asset.disponibile
        })

    return jsonify({
        'categories': categories,
        'total': len(assets)
    }), 200
